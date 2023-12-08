import os
import argparse
from typing import List, Dict, cast
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
from api import BaseApi, GigaChatApi


parser = argparse.ArgumentParser(
    description="Parse files content to read and check that")
parser.add_argument("-c", "--classes", type=Path,
                    help="Path to the file with classes.",
                    required=True)
parser.add_argument("-s", "--strings", type=Path,
                    help="Path to the file to be checked",
                    required=True)
parser.add_argument("-a", "--adds", type=str,
                    help="Your adds for request",
                    required=False)
args = parser.parse_args()
classes = str(args.classes.absolute())
strings = str(args.strings.absolute())
adds = args.adds
print(f"run {classes=} {strings}")


class Tester:
    def __init__(self,
                 api: BaseApi
                 ) -> None:
        self.api = api

    def convert_exc_to_dict(self) -> Dict[int, str]:
        print(f"run load {classes}")
        workbook = load_workbook(classes)
        print(f"load {classes} success")
        sheet = workbook.active
        sheet = cast(Worksheet, workbook.active)
        assert sheet, "sheet is None"
        data_dict: dict = {}

        for row in sheet.iter_rows(  # type: ignore
            min_row=2,
            values_only=True
        ):
            key = row[0]
            value = str(row[1]) if row[1] is not None else ""
            data_dict[key] = value
        return data_dict

    def create_prompt(self,
                      adds: str,
                      line1: str,
                      line2: str
                      ) -> List[Dict[str, str]]:
        print(line1, '=', line2)
        return [
            {
                'role': 'system',
                'content': (
                    'Ты филолог, помогающий пользователю сравнить '
                    'два предложения по его смысловому содержанию. '
                    'Ты проводишь семантический анализ этих предложений и стараешься '
                    'быть максимально точным. Ты отвечаешь только Да и Нет на вопрос: '
                    '"Касается ли тема одного предложения темы другого предложения?" '
                    'либо на вопрос: '
                    '"Раскрывает ли одно предложение смысл другого предложения?" '
                    f'{adds}'
                )
            },
            {
                'role': 'user',
                'content': (
                    'предложение 1: "Мама мыла раму"\n '
                    'предложение 2: "Копать картошку хорошо"'
                )
            },
            {
                'role': 'assistant',
                'content': 'Нет'
            },
            {
                'role': 'user',
                'content': (
                    'предложение 1: "Мама мыла раму"\n'
                    'предложение 2: "Мать устраняла грязь на раме"'
                )
            },
            {
                'role': 'assistant',
                'content': 'Да'
            },
            {
                'role': 'user',
                'content': (
                    f'предложение 1: "{line1}"\n'
                    f'предложение 2: "{line2}"'
                )
            }
        ]

    def run(self) -> None:
        data_dict = self.convert_exc_to_dict()
        # Открытие эксель файла с работой дата-инженера
        print(f"run load {strings}")
        workbook = load_workbook(
            strings,
            data_only=True
        )
        print(f"load {strings} success")
        assert workbook.active, "sheet is None"
        sheet = cast(Worksheet, workbook.active)

        current_row = 2
        # Перебор строк в рабочем Excel файле

        while current_row <= sheet.max_row:

            tagged_id = sheet.cell(
                row=current_row,
                column=1
            ).value
            if tagged_id is None:
                continue
            if isinstance(tagged_id, (float, str)):
                itagged_id = int(tagged_id)
            elif isinstance(tagged_id, int):
                itagged_id = int(tagged_id)
            else:
                continue
            correct_line = data_dict.get(itagged_id)
            checked_line = sheet.cell(
                row=current_row,
                column=2
            ).value
            if (
                (
                    not checked_line or
                    not isinstance(checked_line, str)
                ) or
                (
                    not correct_line or
                    not isinstance(correct_line, str)
                )
            ):
                continue
            rsp = self.api.completions(
                messages=self.create_prompt(
                    adds,
                    correct_line,
                    checked_line
                )
            )
            print(rsp)
            if 'нет' in rsp.lower():
                cell = sheet.cell(
                    row=current_row,
                    column=2
                )
                cell.fill = PatternFill(
                    start_color="FF0000",
                    end_color="FF0000",
                    fill_type='solid'
                )
            elif 'извините' in rsp.lower():
                cell = sheet.cell(
                    row=current_row,
                    column=2
                )
                cell.fill = PatternFill(
                    start_color="FFFF00",
                    end_color="FFFF00",
                    fill_type='solid'
                )

            else:
                cell = sheet.cell(
                    row=current_row,
                    column=2
                )
                cell.fill = PatternFill(
                    start_color="FFFFFF",
                    end_color="FFFFFF",
                    fill_type='solid'
                )
            current_row += 1
        workbook.save(strings)


if __name__ == "__main__":
    AUTORIZATION_DATA = os.getenv('AUTORIZATION_DATA')
    assert AUTORIZATION_DATA, "AUTORIZATION_DATA is None"

    api = GigaChatApi(AUTORIZATION_DATA)
    tester = Tester(api)
    tester.run()
